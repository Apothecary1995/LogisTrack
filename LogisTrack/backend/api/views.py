from collections import defaultdict
from datetime import timedelta
from decimal import Decimal, InvalidOperation

from django.db.models import Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken

from .models import (
    DriverLeave,
    DriverProfile,
    ERPSetting,
    FuelEntry,
    PayrollEntry,
    RouteDistance,
    ServiceRepair,
    Trip,
    User,
    Vehicle,
)
from .permissions import IsCompanyAdminOrReadOnly
from .serializers import (
    ChangePasswordSerializer,
    DateRangeSerializer,
    DriverLeaveSerializer,
    DriverProfileSerializer,
    ERPSettingSerializer,
    ForgotPasswordSerializer,
    FuelEntrySerializer,
    FuelMergedRowSerializer,
    LoginSerializer,
    PayrollEntrySerializer,
    RegisterSerializer,
    RouteDistanceSerializer,
    ServiceRepairSerializer,
    TripSerializer,
    UserSerializer,
    UserPreferenceSerializer,
    VehicleSerializer,
)


def _to_decimal(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    parsed = (
        str(value)
        .strip()
        .replace("km", "")
        .replace("KM", "")
        .replace("₺", "")
        .replace("$", "")
        .replace(",", ".")
    )
    if not parsed:
        return None

    try:
        return Decimal(parsed)
    except InvalidOperation:
        return None


def _period_key(value_date, view_type):
    if view_type == "weekly":
        week_start = value_date - timedelta(days=value_date.weekday())
        return week_start.isoformat()
    return value_date.strftime("%Y-%m")


def _group_amounts(rows, view_type):
    grouped = defaultdict(lambda: Decimal("0"))
    for row_date, row_amount in rows:
        grouped[_period_key(row_date, view_type)] += row_amount or Decimal("0")

    sorted_keys = sorted(grouped.keys(), reverse=True)
    return [{"period": period, "amount": grouped[period]} for period in sorted_keys]


class CompanyScopedQuerysetMixin:
    model = None

    def get_queryset(self):
        user = self.request.user
        base_queryset = self.model.objects.all()

        if user.is_superuser and not user.company_id:
            return base_queryset
        if not user.company_id:
            return self.model.objects.none()
        return base_queryset.filter(company=user.company)


class HealthCheckView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        return Response({"status": "ok", "service": "logitarget-backend"})


class UserCountView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        return Response({"registered_user_count": User.objects.count()})


class RegisterView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = RegisterSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()

        refresh = RefreshToken.for_user(user)
        return Response(
            {
                "user": UserSerializer(user).data,
                "access": str(refresh.access_token),
                "refresh": str(refresh),
            },
            status=status.HTTP_201_CREATED,
        )


class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data["user"]

        refresh = RefreshToken.for_user(user)
        return Response(
            {
                "user": UserSerializer(user).data,
                "access": str(refresh.access_token),
                "refresh": str(refresh),
            }
        )


class ForgotPasswordView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = ForgotPasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        # Placeholder: Email integration is intentionally not wired in this MVP.
        return Response(
            {
                "message": "If the email exists, password reset instructions have been queued.",
            }
        )


class MeView(APIView):
    def get(self, request):
        return Response(UserSerializer(request.user).data)


class ChangePasswordView(APIView):
    def post(self, request):
        serializer = ChangePasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        current_password = serializer.validated_data["current_password"]
        new_password = serializer.validated_data["new_password"]

        if not request.user.check_password(current_password):
            return Response({"detail": "Current password is incorrect."}, status=status.HTTP_400_BAD_REQUEST)

        request.user.set_password(new_password)
        request.user.save(update_fields=["password"])

        refresh = RefreshToken.for_user(request.user)
        return Response(
            {
                "message": "Password updated successfully.",
                "access": str(refresh.access_token),
                "refresh": str(refresh),
            }
        )


class UserPreferenceView(APIView):
    def get(self, request):
        return Response(UserPreferenceSerializer(request.user).data)

    def patch(self, request):
        serializer = UserPreferenceSerializer(request.user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(UserSerializer(request.user).data)


class VehicleViewSet(CompanyScopedQuerysetMixin, viewsets.ModelViewSet):
    model = Vehicle
    serializer_class = VehicleSerializer

    def _ensure_driver_profile(self, vehicle):
        if not vehicle.driver_name:
            return
        DriverProfile.objects.get_or_create(
            company=vehicle.company,
            full_name=vehicle.driver_name.strip(),
            defaults={"is_active": True},
        )

    def perform_create(self, serializer):
        vehicle = serializer.save(company=self.request.user.company)
        self._ensure_driver_profile(vehicle)

    def perform_update(self, serializer):
        vehicle = serializer.save()
        self._ensure_driver_profile(vehicle)

    @action(detail=True, methods=["post"], url_path="create-trip")
    def create_trip(self, request, pk=None):
        vehicle = self.get_object()
        payload = request.data.copy()
        payload["vehicle"] = vehicle.id

        serializer = TripSerializer(data=payload, context={"request": request})
        serializer.is_valid(raise_exception=True)
        trip = serializer.save()

        return Response(TripSerializer(trip, context={"request": request}).data, status=status.HTTP_201_CREATED)


class RouteDistanceViewSet(CompanyScopedQuerysetMixin, viewsets.ModelViewSet):
    model = RouteDistance
    serializer_class = RouteDistanceSerializer
    permission_classes = [IsCompanyAdminOrReadOnly]


class TripViewSet(CompanyScopedQuerysetMixin, viewsets.ModelViewSet):
    model = Trip
    serializer_class = TripSerializer

    def get_queryset(self):
        queryset = super().get_queryset().select_related("vehicle")
        start_date = self.request.query_params.get("start_date")
        end_date = self.request.query_params.get("end_date")

        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)

        return queryset


class ServiceRepairViewSet(CompanyScopedQuerysetMixin, viewsets.ModelViewSet):
    model = ServiceRepair
    serializer_class = ServiceRepairSerializer


class FuelEntryViewSet(CompanyScopedQuerysetMixin, viewsets.ModelViewSet):
    model = FuelEntry
    serializer_class = FuelEntrySerializer


class PayrollEntryViewSet(CompanyScopedQuerysetMixin, viewsets.ModelViewSet):
    model = PayrollEntry
    serializer_class = PayrollEntrySerializer


class DriverProfileViewSet(CompanyScopedQuerysetMixin, viewsets.ModelViewSet):
    model = DriverProfile
    serializer_class = DriverProfileSerializer


class DriverLeaveViewSet(CompanyScopedQuerysetMixin, viewsets.ModelViewSet):
    model = DriverLeave
    serializer_class = DriverLeaveSerializer


class DashboardSummaryView(APIView):
    def get(self, request):
        company = request.user.company
        if not company:
            return Response({"detail": "User is not linked to a company."}, status=status.HTTP_400_BAD_REQUEST)

        today = timezone.localdate()
        month_start = today.replace(day=1)

        trip_queryset = Trip.objects.filter(company=company)
        monthly_revenue = (
            trip_queryset.filter(created_at__date__gte=month_start, created_at__date__lte=today).aggregate(
                total=Coalesce(Sum("total_amount"), Decimal("0"))
            )["total"]
            or Decimal("0")
        )

        total_trips = trip_queryset.count()
        total_components = trip_queryset.aggregate(
            cci=Coalesce(Sum("cci_km"), Decimal("0")),
            extra=Coalesce(Sum("extra_km"), Decimal("0")),
        )
        total_kilometers = (total_components["cci"] or Decimal("0")) + (total_components["extra"] or Decimal("0"))

        pending_maintenance = Vehicle.objects.filter(
            company=company,
            last_inspection_date__lt=today - timedelta(days=365),
        ).count()

        return Response(
            {
                "monthly_revenue": monthly_revenue,
                "pending_maintenance": pending_maintenance,
                "total_trips": total_trips,
                "total_kilometers": total_kilometers,
                "routes_in_database": RouteDistance.objects.filter(company=company).count(),
            }
        )


class RouteDistanceUploadView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [IsCompanyAdminOrReadOnly]

    def _normalize_header(self, header_value):
        return (
            str(header_value or "")
            .strip()
            .lower()
            .replace(" ", "")
            .replace("_", "")
            .replace("-", "")
            .replace("/", "")
        )

    def _parse_route_text(self, route_text):
        raw = str(route_text or "").strip()
        if not raw:
            return None, None

        for separator in ["->", ">", "-", "/", "|"]:
            if separator in raw:
                parts = [part.strip() for part in raw.split(separator) if part.strip()]
                if len(parts) >= 2:
                    return parts[0], parts[1]

        return None, None

    def post(self, request):
        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "Excel file is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = load_workbook(upload, data_only=True)
        except Exception:
            return Response({"detail": "Invalid Excel file. Please upload .xlsx format."}, status=status.HTTP_400_BAD_REQUEST)

        worksheet = workbook.active
        rows = worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return Response({"detail": "Excel header row is missing."}, status=status.HTTP_400_BAD_REQUEST)

        normalized_headers = [self._normalize_header(cell) for cell in header_row]

        def find_index(options):
            for option in options:
                if option in normalized_headers:
                    return normalized_headers.index(option)
            return None

        origin_idx = find_index(["origin", "cikis", "start", "from"])
        destination_idx = find_index(["destination", "varis", "end", "to"])
        route_idx = find_index(["origindestination", "route", "rota", "parkur"])
        km_idx = find_index(["km", "distance", "mesafe", "cci"])

        if km_idx is None:
            return Response({"detail": "KM column not found in uploaded file."}, status=status.HTTP_400_BAD_REQUEST)

        upsert_count = 0
        skipped_count = 0

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            origin = row[origin_idx].strip() if origin_idx is not None and row[origin_idx] else None
            destination = row[destination_idx].strip() if destination_idx is not None and row[destination_idx] else None

            if (not origin or not destination) and route_idx is not None:
                parsed_origin, parsed_destination = self._parse_route_text(row[route_idx])
                origin = origin or parsed_origin
                destination = destination or parsed_destination

            km_value = _to_decimal(row[km_idx]) if km_idx is not None and km_idx < len(row) else None

            if not origin or not destination or km_value is None:
                skipped_count += 1
                continue

            RouteDistance.objects.update_or_create(
                company=request.user.company,
                origin=origin,
                destination=destination,
                defaults={"km": km_value},
            )
            upsert_count += 1

        return Response(
            {
                "message": "Route database synchronized.",
                "upserted": upsert_count,
                "skipped": skipped_count,
            },
            status=status.HTTP_200_OK,
        )


class ArchiveExportView(APIView):
    def get(self, request):
        trips = Trip.objects.filter(company=request.user.company).select_related("vehicle")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Archive"

        headers = [
            "Created At",
            "Plate",
            "Driver",
            "Origin",
            "Destination",
            "Departure",
            "Arrival",
            "Duration",
            "CCI KM",
            "Extra KM",
            "Total KM",
            "Cargo Type",
            "Quantity",
            "Customer",
            "Price",
            "Total Amount",
            "Waybill No",
            "Invoice No",
            "Invoice Date",
            "Notes",
        ]
        worksheet.append(headers)

        for trip in trips:
            total_km = (trip.cci_km or Decimal("0")) + (trip.extra_km or Decimal("0"))
            worksheet.append(
                [
                    trip.created_at.strftime("%Y-%m-%d %H:%M"),
                    trip.plate_number,
                    trip.vehicle.driver_name,
                    trip.origin,
                    trip.destination,
                    trip.departure_time.strftime("%Y-%m-%d %H:%M") if trip.departure_time else "",
                    trip.arrival_time.strftime("%Y-%m-%d %H:%M") if trip.arrival_time else "",
                    trip.total_duration,
                    float(trip.cci_km or 0),
                    float(trip.extra_km or 0),
                    float(total_km),
                    trip.cargo_type,
                    float(trip.quantity or 0),
                    trip.customer,
                    float(trip.price or 0),
                    float(trip.total_amount or 0),
                    trip.waybill_no,
                    trip.invoice_no,
                    trip.invoice_date.isoformat() if trip.invoice_date else "",
                    trip.notes,
                ]
            )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"fleet-archive-{timezone.localdate().isoformat()}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response


class FuelMergedView(APIView):
    def get(self, request):
        company_entries = FuelEntry.objects.filter(company=request.user.company).select_related("vehicle")

        merged = defaultdict(
            lambda: {
                "date": None,
                "plate_number": "",
                "fuel_liters": Decimal("0"),
                "fuel_amount": Decimal("0"),
                "adblue_liters": Decimal("0"),
                "adblue_amount": Decimal("0"),
                "notes": "",
            }
        )

        for entry in company_entries:
            key = (entry.date, entry.vehicle.plate_number)
            row = merged[key]
            row["date"] = entry.date
            row["plate_number"] = entry.vehicle.plate_number

            if entry.entry_type == FuelEntry.FUEL:
                row["fuel_liters"] += entry.liters
                row["fuel_amount"] += entry.amount
            else:
                row["adblue_liters"] += entry.liters
                row["adblue_amount"] += entry.amount

            if entry.notes:
                row["notes"] = f"{row['notes']} | {entry.notes}".strip(" |")

        rows = sorted(merged.values(), key=lambda item: (item["date"], item["plate_number"]), reverse=True)
        serializer = FuelMergedRowSerializer(rows, many=True)
        return Response(serializer.data)


class PayrollOverviewView(APIView):
    def get(self, request):
        params_serializer = DateRangeSerializer(data=request.query_params)
        params_serializer.is_valid(raise_exception=True)
        params = params_serializer.validated_data

        today = timezone.localdate()
        start_date = params.get("start_date") or today.replace(day=1)
        end_date = params.get("end_date") or today
        view_type = params.get("view", "monthly")

        company = request.user.company

        trips = Trip.objects.filter(company=company)
        service_repairs = ServiceRepair.objects.filter(company=company, date__range=(start_date, end_date))
        fuel_entries = FuelEntry.objects.filter(company=company, date__range=(start_date, end_date))
        payroll_entries = PayrollEntry.objects.filter(company=company, date__range=(start_date, end_date))

        trip_invoice_rows = []
        for trip in trips:
            invoice_reference_date = trip.invoice_date or trip.created_at.date()
            if start_date <= invoice_reference_date <= end_date:
                trip_invoice_rows.append((invoice_reference_date, trip.total_amount or Decimal("0")))

        wage_rows = [
            (entry.date, entry.amount or Decimal("0"))
            for entry in payroll_entries
            if entry.entry_type == PayrollEntry.DRIVER_WAGE
        ]

        customer_invoice_rows = trip_invoice_rows + [
            (entry.date, entry.amount or Decimal("0"))
            for entry in payroll_entries
            if entry.entry_type == PayrollEntry.CUSTOMER_INVOICE
        ]

        trip_income_total = sum([amount for _, amount in trip_invoice_rows], Decimal("0"))
        manual_income_total = payroll_entries.filter(
            entry_type__in=[PayrollEntry.CUSTOMER_INVOICE, PayrollEntry.OTHER_INCOME]
        ).aggregate(total=Coalesce(Sum("amount"), Decimal("0")))["total"]

        manual_expense_total = payroll_entries.filter(
            entry_type__in=[PayrollEntry.DRIVER_WAGE, PayrollEntry.OTHER_EXPENSE]
        ).aggregate(total=Coalesce(Sum("amount"), Decimal("0")))["total"]

        service_total = service_repairs.aggregate(total=Coalesce(Sum("cost"), Decimal("0")))["total"]
        fuel_total = fuel_entries.aggregate(total=Coalesce(Sum("amount"), Decimal("0")))["total"]

        net_balance = (trip_income_total + manual_income_total) - (service_total + fuel_total + manual_expense_total)

        return Response(
            {
                "range": {
                    "start_date": start_date,
                    "end_date": end_date,
                    "view": view_type,
                },
                "totals": {
                    "trip_income": trip_income_total,
                    "manual_income": manual_income_total,
                    "service_expense": service_total,
                    "fuel_expense": fuel_total,
                    "manual_expense": manual_expense_total,
                    "net_balance": net_balance,
                },
                "driver_wages": _group_amounts(wage_rows, view_type),
                "customer_invoices": _group_amounts(customer_invoice_rows, view_type),
            }
        )


class EmployeeInsightView(APIView):
    def get(self, request):
        company = request.user.company
        metric = request.query_params.get("metric", "most_km")

        stats = defaultdict(
            lambda: {
                "driver_name": "",
                "total_km": Decimal("0"),
                "total_amount": Decimal("0"),
                "efficiency": Decimal("0"),
            }
        )

        drivers = DriverProfile.objects.filter(company=company)
        for driver in drivers:
            stats[driver.full_name]["driver_name"] = driver.full_name

        trips = Trip.objects.filter(company=company).select_related("vehicle")
        for trip in trips:
            driver_name = (trip.vehicle.driver_name or "Unassigned").strip()
            row = stats[driver_name]
            row["driver_name"] = driver_name
            row["total_km"] += (trip.cci_km or Decimal("0")) + (trip.extra_km or Decimal("0"))
            row["total_amount"] += trip.total_amount or Decimal("0")

        for value in stats.values():
            if value["total_km"] > 0:
                value["efficiency"] = value["total_amount"] / value["total_km"]

        rows = list(stats.values())

        if metric == "least_km":
            rows.sort(key=lambda item: item["total_km"])
        elif metric == "most_efficient":
            rows.sort(key=lambda item: item["efficiency"], reverse=True)
        elif metric == "least_efficient":
            rows.sort(key=lambda item: item["efficiency"])
        else:
            rows.sort(key=lambda item: item["total_km"], reverse=True)

        leaves = DriverLeave.objects.filter(company=company).select_related("driver")

        return Response(
            {
                "metric": metric,
                "drivers": rows,
                "leaves": DriverLeaveSerializer(leaves, many=True).data,
            }
        )


class ERPSettingView(APIView):
    permission_classes = [IsCompanyAdminOrReadOnly]

    def get_object(self, user):
        setting, _ = ERPSetting.objects.get_or_create(company=user.company)
        return setting

    def get(self, request):
        setting = self.get_object(request.user)
        return Response(ERPSettingSerializer(setting).data)

    def put(self, request):
        setting = self.get_object(request.user)
        serializer = ERPSettingSerializer(setting, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def patch(self, request):
        setting = self.get_object(request.user)
        serializer = ERPSettingSerializer(setting, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)
